# -*- coding: utf-8 -*-
"""
LootBar 荒野行動マーケット 定期取得パイプライン
=====================================================
GitHub Actionsから定期実行される想定の統合スクリプト。

やること:
  1. 為替レート(USD->JPY)を無料APIから取得
  2. 一覧API・タグAPIを取得
  3. 各商品の価格トレンド(過去30日)を取得
  4. 商品名・ジャンルを日本語化(translate_dict.py)
  5. 本日分のスナップショットを data/history/YYYY-MM-DD.csv に保存
  6. 直前のスナップショットと比較し、価格変動が大きい商品を検出
  7. 要件通りの8列+検証用列のExcelを output/lootbar_knivesout_market_list.xlsx に生成
  8. 変動サマリー付きメールを送信(Excel添付)

環境変数(GitHub Secretsから渡される想定):
  SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASSWORD, MAIL_TO
"""

import csv
import glob
import os
import re
import smtplib
import time
import random
from datetime import datetime, timezone
from email.message import EmailMessage
from pathlib import Path

import requests
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from translate_dict import translate_genre, translate_name

# ---------- 設定 ----------
GAME = "knivesout"
SUBGAME = "10007"
PAGE_SIZE = 40
DAYS = 30
RETRY = 3
SLEEP_MIN = 0.5
SLEEP_MAX = 1.0
PRICE_CHANGE_ALERT_THRESHOLD = 0.15  # 15%以上の変動で通知対象

BASE = "https://api.lootbar.com"
LIST_URL = f"{BASE}/api/market/goods"
TAGS_URL = f"{BASE}/api/market/tags"
PRICE_HISTORY_URL = f"{BASE}/api/market/goods/price_history/"

# 調査の結果判明したLootBar独自ヘッダー。これを付けるとAPIが商品名・
# ジャンルを日本語で返してくる(Cookie/URL/クエリパラメータではなく
# このヘッダーが決め手だった)。
# 注意: 実際のブラウザは同時に x-currency: JPY も送っていたが、
# 価格そのものが円建てに変換されて返る可能性があり、既存の
# 自前USD->JPY換算ロジックと二重変換になるリスクがあるため、
# あえて locale のみを付与する(価格系のAPIレスポンス形式は
# 変えず、テキスト翻訳の恩恵だけを受ける)。
JP_HEADERS = {
    "x-ps-locale": "ja",
}

# 日本語判定用(ひらがな・カタカナ・漢字のいずれかを含むか)
JP_CHAR_PATTERN = re.compile(r"[\u3040-\u30ff\u4e00-\u9fff]")

ROOT = Path(__file__).resolve().parent.parent
HISTORY_DIR = ROOT / "data" / "history"
OUTPUT_DIR = ROOT / "output"
HISTORY_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

TODAY = datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d")
SNAPSHOT_CSV = HISTORY_DIR / f"{TODAY}.csv"
OUT_XLSX = OUTPUT_DIR / "lootbar_knivesout_market_list.xlsx"

log_lines = []


def log(msg: str):
    print(msg)
    log_lines.append(msg)


def sleep_polite():
    time.sleep(random.uniform(SLEEP_MIN, SLEEP_MAX))


# ---------- 為替レート取得 ----------
def fetch_usd_jpy_rate() -> tuple[float, str]:
    """無料APIから最新のUSD->JPYレートを取得。失敗時は固定値にフォールバック。"""
    try:
        resp = requests.get("https://open.er-api.com/v6/latest/USD", timeout=10)
        resp.raise_for_status()
        data = resp.json()
        rate = float(data["rates"]["JPY"])
        source = f"open.er-api.com (取得日時: {data.get('time_last_update_utc', TODAY)})"
        log(f"[為替レート] 1USD = {rate}円 ({source})")
        return rate, source
    except Exception as e:
        fallback = 150.0
        log(f"[警告] 為替レート取得失敗({e})。フォールバック値 {fallback} を使用します。")
        return fallback, "取得失敗のためフォールバック固定値を使用"


# ---------- タグ(ジャンル)マップ ----------
def build_genre_internal_map(req, extra_headers=None) -> dict:
    resp = req.get(f"{TAGS_URL}?game={GAME}", headers=extra_headers or {})
    body = resp.json()
    m = {}

    def walk(node):
        value = node.get("value")
        title = node.get("title")
        if value:
            m[value] = title
        for c in node.get("children", []):
            walk(c)

    for item in body["data"]["items"]:
        walk(item)
    return m


# ---------- 一覧取得 ----------
def fetch_all_list_items(req, extra_headers=None, label=""):
    tag = f"[{label}]" if label else ""
    all_items = {}
    page_num = 1
    total_page = 1
    declared_total_count = None

    while page_num <= total_page:
        url = f"{LIST_URL}?game={GAME}&subgame_name={SUBGAME}&page_num={page_num}&page_size={PAGE_SIZE}"
        body = None
        for attempt in range(1, RETRY + 1):
            try:
                resp = req.get(
                    url,
                    headers={
                        "Referer": "https://www.lootbar.com/ja/market/all",
                        **(extra_headers or {}),
                    },
                )
                if resp.ok:
                    candidate = resp.json()
                    if candidate.get("status") == "ok":
                        body = candidate
                        break
            except Exception as e:
                log(f"[リトライ {attempt}/{RETRY}] page {page_num} 例外: {e}")
            sleep_polite()

        if body is None:
            log(f"[エラー] page {page_num} を{RETRY}回試行しても取得失敗。スキップします。")
            page_num += 1
            continue

        data = body["data"]
        total_page = data["total_page"]
        declared_total_count = data["total_count"]

        for it in data["items"]:
            gid = it["goods_id"]
            all_items[gid] = it

        log(f"{tag}[一覧取得] page {page_num}/{total_page}  累計{len(all_items)}件")
        page_num += 1
        sleep_polite()

    log(f"{tag}[一覧取得完了] 取得件数={len(all_items)}  API申告件数={declared_total_count}")
    if declared_total_count is not None and len(all_items) != declared_total_count:
        log(f"{tag}[!!! 不一致警告 !!!] 取得件数とAPI申告件数が一致しません")

    return all_items, declared_total_count


def fetch_price_history(req, goods_id):
    url = f"{PRICE_HISTORY_URL}?game={GAME}&goods_id={goods_id}&days={DAYS}&subgame_name={SUBGAME}"
    for attempt in range(1, RETRY + 1):
        try:
            resp = req.get(
                url,
                headers={"Referer": f"https://www.lootbar.com/ja/market/goods/{goods_id}"},
            )
            if resp.ok:
                body = resp.json()
                if body.get("status") == "ok":
                    return body["data"]["price_history"]
        except Exception as e:
            log(f"    [リトライ {attempt}/{RETRY}] goods_id={goods_id} 例外: {e}")
        sleep_polite()
    return None


# ---------- スクレイピング本体 ----------
def scrape(rate: float):
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="ja-JP",
        )
        page = context.new_page()
        page.goto(
            f"https://www.lootbar.com/ja/market/all?game={GAME}&subgame_name={SUBGAME}",
            wait_until="networkidle",
            timeout=60000,
        )
        time.sleep(2)
        req = context.request

        # 英語版・日本語版をそれぞれ取得(goods_idで突き合わせる)。
        # 日本語版は x-ps-locale: ja ヘッダーにより、APIが商品名・
        # ジャンルをあらかじめ日本語で返してくる(調査で判明した挙動)。
        genre_internal_map_en = build_genre_internal_map(req)
        genre_internal_map_ja = build_genre_internal_map(req, extra_headers=JP_HEADERS)

        all_items_en, declared_total_count = fetch_all_list_items(req, label="EN")
        all_items_ja, _ = fetch_all_list_items(req, extra_headers=JP_HEADERS, label="JA")

        records = []
        for i, (gid, it) in enumerate(sorted(all_items_en.items(), key=lambda x: x[0]), start=1):
            name_en = it.get("name", "")
            goods_info = it.get("goods_info", {}) or {}
            tags = goods_info.get("tags", {}) or {}
            fc_internal = (tags.get("first_class") or {}).get("internal_name", "")
            sc_internal = (tags.get("second_class") or {}).get("internal_name", "")
            fc_title_en = genre_internal_map_en.get(fc_internal, "")
            sc_title_en = genre_internal_map_en.get(sc_internal, "")
            genre_en = f"{fc_title_en} / {sc_title_en}" if sc_title_en else fc_title_en

            it_ja = all_items_ja.get(gid)
            name_from_api_ja = (it_ja or {}).get("name", "")
            fc_title_ja = genre_internal_map_ja.get(fc_internal, "")
            sc_title_ja = genre_internal_map_ja.get(sc_internal, "")
            genre_from_api_ja = f"{fc_title_ja} / {sc_title_ja}" if sc_title_ja else fc_title_ja

            has_leftover = False

            # 商品名: APIが日本語(ひらがな/カタカナ/漢字)を返していればそれを採用。
            # 返していなければ(新商品でAPI側の翻訳データがまだ無い等)、
            # 従来の辞書ベース翻訳をフォールバックとして使う。
            if JP_CHAR_PATTERN.search(name_from_api_ja):
                name_ja = name_from_api_ja
            else:
                name_ja, _ = translate_name(name_en)
                has_leftover = True  # API側が未対応だったこと自体を要確認扱いにする

            # ジャンル: 同様にAPI側の日本語タイトルを優先し、
            # 無ければ辞書(GENRE_MAP)にフォールバック。
            if genre_from_api_ja and JP_CHAR_PATTERN.search(genre_from_api_ja):
                genre_ja = genre_from_api_ja
            else:
                genre_ja = translate_genre(genre_en)

            price_hist = fetch_price_history(req, gid)
            last_price_jpy = ""
            last_date = ""
            count_30d = ""
            avg_price_jpy = ""
            note = ""

            if price_hist is None:
                note = "価格トレンド取得失敗（要再取得）"
            elif len(price_hist) == 0:
                note = "直近30日間の取引データなし"
            else:
                sorted_hist = sorted(price_hist, key=lambda x: x[0])
                last_ts, last_price_usd = sorted_hist[-1]
                last_price_jpy = round(float(last_price_usd) * rate)
                last_date = time.strftime("%Y-%m-%d", time.gmtime(last_ts))
                count_30d = len(sorted_hist)
                avg_price_usd = sum(float(p) for _, p in sorted_hist) / count_30d
                avg_price_jpy = round(avg_price_usd * rate)
                note = "※価格トレンドAPI(日次データ)に基づく近似値"

            if has_leftover:
                note = (note + " / APIが日本語名を返さず辞書フォールバックを使用(要確認)").strip(" /")

            def to_jpy(usd_str):
                try:
                    return round(float(usd_str) * rate)
                except (TypeError, ValueError):
                    return ""

            records.append({
                "goods_id": gid,
                "商品名": name_ja,
                "商品名_原文": name_en,
                "ジャンル": genre_ja,
                "ジャンル_原文": genre_en,
                "直近成約価格_円": last_price_jpy,
                "直近成約日": last_date,
                "直近30日間の成約数_近似": count_30d,
                "直近30日間の平均成約価格_円": avg_price_jpy,
                "現在出品最安値_円": to_jpy(it.get("sell_min_price")),
                "現在出品最高値_円": to_jpy(it.get("sell_max_price")),
                "現在出品参考価格_円": to_jpy(it.get("sell_reference_price")),
                "現在出品数": it.get("sell_num", ""),
                "備考": note,
                "取得日": TODAY,
            })

            if i % 50 == 0 or i == len(all_items_en):
                log(f"[進捗] {i}/{len(all_items_en)} 件処理完了")
            sleep_polite()

        browser.close()
        return records, declared_total_count


# ---------- スナップショット保存 & 前回比較 ----------
LATEST_CSV = ROOT / "data" / "latest.csv"


def save_snapshot(records):
    fieldnames = list(records[0].keys())
    with open(SNAPSHOT_CSV, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)
    log(f"[保存] スナップショット: {SNAPSHOT_CSV}")

    # GitHub Pages側は日付付きファイル名を追えないため、常に同じ
    # ファイル名(data/latest.csv)でも同内容を書き出す(Web表示用)。
    with open(LATEST_CSV, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)
    log(f"[保存] Web表示用最新データ: {LATEST_CSV}")


def load_previous_snapshot():
    files = sorted(glob.glob(str(HISTORY_DIR / "*.csv")))
    files = [f for f in files if not f.endswith(f"{TODAY}.csv")]
    if not files:
        return None
    prev_path = files[-1]
    with open(prev_path, encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    log(f"[比較対象] 前回スナップショット: {prev_path} ({len(rows)}件)")
    return {r["goods_id"]: r for r in rows}


def detect_price_changes(records, previous_map):
    if not previous_map:
        return []
    changes = []
    for r in records:
        gid = str(r["goods_id"])
        prev = previous_map.get(gid)
        if not prev:
            continue
        try:
            new_avg = float(r["直近30日間の平均成約価格_円"])
            old_avg = float(prev["直近30日間の平均成約価格_円"])
        except (TypeError, ValueError):
            continue
        if old_avg <= 0:
            continue
        pct_change = (new_avg - old_avg) / old_avg
        if abs(pct_change) >= PRICE_CHANGE_ALERT_THRESHOLD:
            changes.append({
                "商品名": r["商品名"],
                "旧平均価格": round(old_avg),
                "新平均価格": round(new_avg),
                "変動率": pct_change,
            })
    changes.sort(key=lambda x: abs(x["変動率"]), reverse=True)
    return changes


# ---------- Excel生成 ----------
FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
NORMAL_FONT = Font(name=FONT_NAME, size=10)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def build_excel(records, rate, rate_source):
    wb = Workbook()
    ws = wb.active
    ws.title = "商品一覧"

    headers = [
        "No.", "商品名", "ジャンル", "直近成約価格（円）", "直近成約日",
        "直近30日間の成約数", "直近30日間の平均成約価格（円）", "備考",
        "goods_id（検証用）", "商品名（原文/英語）", "ジャンル（原文）",
        "現在出品最安値（円）", "現在出品最高値（円）", "現在出品参考価格（円）", "現在出品数",
        "取得日",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    for i, r in enumerate(records, start=2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=r["商品名"])
        ws.cell(row=i, column=3, value=r["ジャンル"])
        ws.cell(row=i, column=4, value=r["直近成約価格_円"] or "")
        ws.cell(row=i, column=5, value=r["直近成約日"])
        ws.cell(row=i, column=6, value=r["直近30日間の成約数_近似"] or "")
        ws.cell(row=i, column=7, value=r["直近30日間の平均成約価格_円"] or "")
        ws.cell(row=i, column=8, value=r["備考"])
        ws.cell(row=i, column=9, value=int(r["goods_id"]))
        ws.cell(row=i, column=10, value=r["商品名_原文"])
        ws.cell(row=i, column=11, value=r["ジャンル_原文"])
        ws.cell(row=i, column=12, value=r["現在出品最安値_円"] or "")
        ws.cell(row=i, column=13, value=r["現在出品最高値_円"] or "")
        ws.cell(row=i, column=14, value=r["現在出品参考価格_円"] or "")
        ws.cell(row=i, column=15, value=r["現在出品数"] or "")
        ws.cell(row=i, column=16, value=r["取得日"])
        for col in (4, 7, 12, 13, 14):
            ws.cell(row=i, column=col).number_format = "#,##0"
        for c in range(1, len(headers) + 1):
            ws.cell(row=i, column=c).font = NORMAL_FONT
            ws.cell(row=i, column=c).border = BORDER

    last_row = len(records) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"
    ws.freeze_panes = "A2"
    widths = [6, 42, 20, 16, 14, 14, 18, 40, 12, 42, 24, 16, 16, 16, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws_note = wb.create_sheet("注記")
    ws_note["A1"] = "為替レート"
    ws_note["B1"] = rate
    ws_note["A2"] = "レート出典"
    ws_note["B2"] = rate_source
    ws_note["A3"] = "取得日"
    ws_note["B3"] = TODAY
    ws_note["A5"] = (
        "商品名は独自の日本語意訳です(LootBar公式訳ではありません)。"
        "武器コードはそのまま残し、スキン名・エディション名のみ翻訳しています。"
        "正確性確認のため「商品名（原文/英語）」列を必ず参照してください。"
    )
    for r in range(1, 4):
        ws_note.cell(row=r, column=1).font = Font(name=FONT_NAME, bold=True, size=10)
    ws_note.column_dimensions["A"].width = 20
    ws_note.column_dimensions["B"].width = 80

    wb.save(OUT_XLSX)
    log(f"[保存] Excel: {OUT_XLSX}")


# ---------- メール送信 ----------
def send_email(records_count, changes, declared_total_count, actual_count):
    host = os.environ.get("SMTP_HOST")
    port = int(os.environ.get("SMTP_PORT", "587"))
    user = os.environ.get("SMTP_USER")
    password = os.environ.get("SMTP_PASSWORD")
    mail_to = os.environ.get("MAIL_TO")

    if not all([host, user, password, mail_to]):
        log("[警告] メール送信用の環境変数が不足しているため、メール送信をスキップします。")
        return

    # MAIL_TOはカンマ区切りで複数指定可能。BCC方式で送るため、
    # Toヘッダーには送信者自身のアドレスを入れ、実際の宛先はSMTP送信時にのみ指定する
    # （メール本文のヘッダーには他の受信者アドレスが一切表示されない）
    recipients = [addr.strip() for addr in mail_to.split(",") if addr.strip()]

    msg = EmailMessage()
    msg["Subject"] = f"[LootBar自動取得] {TODAY} 実行結果（{records_count}件）"
    msg["From"] = user
    msg["To"] = user

    body_lines = [
        f"LootBar 荒野行動マーケットの自動取得が完了しました。",
        f"",
        f"実行日: {TODAY}",
        f"取得件数: {actual_count} / API申告件数: {declared_total_count}",
        f"",
    ]
    if declared_total_count is not None and actual_count != declared_total_count:
        body_lines.append("⚠️ 取得件数がAPI申告件数と一致していません。ログをご確認ください。")
        body_lines.append("")

    if changes:
        body_lines.append(f"■ 前回比15%以上の価格変動があった商品（上位{min(20,len(changes))}件）")
        for c in changes[:20]:
            direction = "↑" if c["変動率"] > 0 else "↓"
            body_lines.append(
                f"  {direction} {c['商品名']}: ¥{c['旧平均価格']:,} → ¥{c['新平均価格']:,} ({c['変動率']*100:+.1f}%)"
            )
    else:
        body_lines.append("■ 前回比で目立った価格変動(±15%以上)はありませんでした。")

    body_lines.append("")
    body_lines.append("詳細は添付のExcelファイルをご確認ください。")

    msg.set_content("\n".join(body_lines))

    with open(OUT_XLSX, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=OUT_XLSX.name,
        )

    with smtplib.SMTP(host, port) as server:
        server.starttls()
        server.login(user, password)
        server.send_message(msg, from_addr=user, to_addrs=recipients)

    log(f"[送信] メール送信完了(BCC方式): {len(recipients)}件の宛先")


# ---------- メイン ----------
def main():
    rate, rate_source = fetch_usd_jpy_rate()
    records, declared_total_count = scrape(rate)
    save_snapshot(records)

    previous_map = load_previous_snapshot()
    changes = detect_price_changes(records, previous_map)

    build_excel(records, rate, rate_source)
    send_email(len(records), changes, declared_total_count, len(records))

    with open(ROOT / "scrape_log.txt", "w", encoding="utf-8") as f:
        f.write("\n".join(log_lines))


if __name__ == "__main__":
    main()